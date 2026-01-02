import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from xhtml2pdf import pisa
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.utils.dateparse import parse_date
from django.utils import timezone
from django.template.loader import render_to_string
from .models import RotinaBackup, ValidacaoBackup
from apps.clientes.models import Cliente, Servidor
from .forms import ValidacaoForm

@login_required
def dashboard(request):
    clientes = Cliente.objects.filter(ativo=True).order_by('nome_fantasia').prefetch_related('servidores', 'rotinas')
    
    for cliente in clientes:
        ultima_validacao = ValidacaoBackup.objects.filter(
            rotina__cliente=cliente
        ).order_by('-created_at').first()
        
        cliente.ultimo_status = ultima_validacao.status if ultima_validacao else 'PENDENTE'
        
        cliente.historico_recente = ValidacaoBackup.objects.filter(
            rotina__cliente=cliente
        ).select_related('rotina', 'rotina__ferramenta', 'usuario').order_by('-created_at')[:5]

    ultimas_validacoes = ValidacaoBackup.objects.select_related(
        'rotina', 'rotina__ferramenta', 'rotina__cliente'
    ).order_by('-created_at')[:10]
    
    return render(request, 'dashboard.html', {
        'clientes': clientes,
        'ultimas_validacoes': ultimas_validacoes
    })

@login_required
def historico_global(request):
    validacoes = ValidacaoBackup.objects.all().select_related(
        'rotina', 'rotina__ferramenta', 'rotina__cliente', 'usuario'
    )

    cliente_id = request.GET.get('cliente')
    status = request.GET.get('status')
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    ordenacao = request.GET.get('ordenacao', 'recente')

    if cliente_id:
        validacoes = validacoes.filter(rotina__cliente_id=cliente_id)
    if status:
        validacoes = validacoes.filter(status=status)
    if data_inicio:
        data_i = parse_date(data_inicio)
        if data_i:
            validacoes = validacoes.filter(created_at__date__gte=data_i)
    if data_fim:
        data_f = parse_date(data_fim)
        if data_f:
            validacoes = validacoes.filter(created_at__date__lte=data_f)

    mapa_ordenacao = {
        'recente': '-created_at',
        'antigo': 'created_at',
        'cliente_az': 'rotina__cliente__nome_fantasia',
        'cliente_za': '-rotina__cliente__nome_fantasia',
        'status': 'status',
    }
    campo_ordem = mapa_ordenacao.get(ordenacao, '-created_at')
    validacoes = validacoes.order_by(campo_ordem)

    clientes = Cliente.objects.filter(ativo=True).order_by('nome_fantasia')
    status_choices = ValidacaoBackup.STATUS_CHOICES

    context = {
        'validacoes': validacoes,
        'clientes': clientes,
        'status_choices': status_choices,
        'total_registros': validacoes.count(),
        'filtros_atuais': request.GET
    }
    return render(request, 'historico_global.html', context)

@login_required
def nova_validacao(request, cliente_id):
    cliente = get_object_or_404(Cliente, pk=cliente_id)
    
    if request.method == 'POST':
        form = ValidacaoForm(cliente_id, request.POST, request.FILES)
        if form.is_valid():
            validacao = form.save(commit=False)
            validacao.usuario = request.user
            validacao.save()
            return redirect('dashboard')
    else:
        form = ValidacaoForm(cliente_id=cliente_id)

    rotinas = RotinaBackup.objects.filter(cliente=cliente)

    return render(request, 'nova_validacao.html', {
        'cliente': cliente,
        'form': form,      
        'rotinas': rotinas
    })

# --- APIs JSON ---
@login_required
def get_servidores_por_cliente(request):
    cliente_id = request.GET.get('cliente_id')
    if not cliente_id:
        return JsonResponse({'servidores': []})
    servidores = Servidor.objects.filter(cliente_id=cliente_id).values('id', 'hostname', 'descricao')
    return JsonResponse({'servidores': list(servidores)})

@login_required
def get_rotinas_cliente(request, cliente_id):
    rotinas = RotinaBackup.objects.filter(cliente_id=cliente_id).values('id', 'descricao', 'ferramenta__nome')
    return JsonResponse({'rotinas': list(rotinas)})

# --- RELATÓRIOS (NOVO CÓDIGO) ---

def _get_queryset_filtrado(request):
    queryset = ValidacaoBackup.objects.all().select_related(
        'rotina', 'rotina__ferramenta', 'rotina__cliente', 'usuario'
    ).order_by('-created_at')

    cliente_id = request.GET.get('cliente')
    status = request.GET.get('status')
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')

    if cliente_id:
        queryset = queryset.filter(rotina__cliente_id=cliente_id)
    if status:
        queryset = queryset.filter(status=status)
    if data_inicio:
        data_i = parse_date(data_inicio)
        if data_i:
            queryset = queryset.filter(created_at__date__gte=data_i)
    if data_fim:
        data_f = parse_date(data_fim)
        if data_f:
            queryset = queryset.filter(created_at__date__lte=data_f)
    return queryset

@login_required
def painel_relatorios(request):
    validacoes = _get_queryset_filtrado(request)
    clientes = Cliente.objects.filter(ativo=True).order_by('nome_fantasia')
    status_choices = ValidacaoBackup.STATUS_CHOICES
    
    context = {
        'validacoes': validacoes[:50],
        'total_registros': validacoes.count(),
        'clientes': clientes,
        'status_choices': status_choices,
    }
    return render(request, 'painel_relatorios.html', context)

@login_required
def gerar_relatorio_excel(request):
    validacoes = _get_queryset_filtrado(request)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Relatorio_Backups_{timezone.now().strftime('%d-%m-%Y')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório de Backups"

    headers = ["Data/Hora", "Cliente", "Servidor", "Ferramenta", "Rotina", "Status", "Usuário"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for item in validacoes:
        servidor = item.rotina.servidores.first()
        cliente_nome = servidor.cliente.nome_fantasia if servidor else "-"
        hostname = servidor.hostname if servidor else "-"
        
        row = [
            item.created_at.strftime("%d/%m/%Y %H:%M"),
            cliente_nome,
            hostname,
            item.rotina.ferramenta.nome,
            item.rotina.descricao,
            item.get_status_display(),
            item.usuario.username
        ]
        ws.append(row)

    for column_cells in ws.columns:
        length = max(len(str(cell.value) or "") for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    wb.save(response)
    return response

@login_required
def gerar_relatorio_pdf(request):
    validacoes = _get_queryset_filtrado(request)
    
    context = {
        'validacoes': validacoes,
        'usuario': request.user.username,
        'data_geracao': timezone.now()
    }
    
    html_string = render_to_string('relatorio_pdf_template.html', context)
    
    response = HttpResponse(content_type='application/pdf')
    filename = f"Relatorio_Backups_{timezone.now().strftime('%d-%m-%Y')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    pisa_status = pisa.CreatePDF(html_string, dest=response)
    
    if pisa_status.err:
        return HttpResponse('Erro ao gerar PDF', status=500)
        
    return response